import re
from openpyxl import Workbook
import Porter_Stemmer_Python

class SentenceProcessor:
    def __init__(self):
        self.sentences = open("sentences.txt", "r", encoding="utf8")  # sets up file to read from
        self.processed = open("processed.txt", "w+", encoding="utf8")  # creates file to hold processed sentences
        self.stop_words = open("stop_words.txt", "r", encoding="utf8")
        self.stop = []
        self.lines = []
        self.create_stop_words()
        self.words = [[]]
        self.words[0].insert(0, "Keyword Set")
        self.book = Workbook()
        self.reduced = Workbook()

        self.process_sentences()

        self.book.save("tdm.xlsx")
        self.reduced.save("reduced_tdm.xlsx")

    def remove_characters(self):
        line = self.current
        for char in '0123456789`~!@#$%^&*()-_=+[{]};:,<.>/?\'"\n':
            line = line.replace(char, '')
        return line

    def create_stop_words(self):
        for i in range(119):
            word = self.stop_words.readline().replace("\n", "")
            word = " " + word + " "
            self.stop.append(word)

    def remove_stop_words(self):
        line = self.current
        for word in self.stop:
            line = line.replace(word, ' ')
        return line

    def add_to_words(self, word):
        if word in self.words[0]:
            return
        else:
            self.words[0].insert(len(self.words[0]), word)

    def add_sentence_to_words(self):
        for k in range(46):
            line = self.processed.readline()
            sentence = line.split(' ')
            sen = [0] * len(self.words[0])
            sen.insert(0, ''.join("Sentence " + str(k + 1)))
            for i in range(len(self.words[0]) - 1):
                count = 0
                for j in range(len(sentence)):
                    if sentence[j] == self.words[0][i + 1]:
                        count = count + 1
                sen[i + 1] = count
                # sen.insert(i + 1, count)
            # print(sen)
            self.words.insert(k + 1, sen)

    def create_excel(self):
        sheet = self.book.active
        for i in range(len(self.words)):
            for j in range(len(self.words[0])):
                # print(str(i) + " :" + str(j))
                sheet.cell(row=(i + 1), column=(j + 1)).value = self.words[i][j]

    def create_reduced(self):
        sheet = self.reduced.active
        for i in range(len(self.words)):
            for j in range(len(self.words[0])):
                # print(str(i) + " :" + str(j))
                sheet.cell(row=(i + 1), column=(j + 1)).value = self.words[i][j]

    def reduce_tdm(self):
        print(self.words)
        columns = []
        columns.append("Totals")
        for i in range(len(self.words[0]) - 1):
            count = 0
            for j in range(len(self.words) - 1):
                if self.words[j + 1][i + 1] > 0:
                    count = count + self.words[j + 1][i + 1]
            columns.append(count)
        # print(columns)
        threshold = 3
        i = len(columns) - 1
        while i >= 1:
            if columns[i] < threshold:
                for j in range(len(self.words) - 1):
                    del self.words[j][i]
            i = i - 1
        print(self.words)

    def process_sentences(self):
        for i in range(46):
            self.current = " " + self.sentences.readline()  # sets current to a line in sentences.txt
            self.current = self.remove_characters()  # removes numbers and special characters
            self.current = self.current.lower()  # makes all characters lowercase
            self.current = self.remove_stop_words()  # removes stop words from sentences
            self.current = re.sub(' +', ' ', self.current)  # trims excess spaces
            self.current = self.current.split(' ')  # splits the sentence into an array of words
            self.current.pop(0)  # removes the first element because it is ''
            self.current.pop(len(self.current) - 1)  # removes the last element because its is ''
            # print(self.current)  # this is for testing
            for j in range(len(self.current)):
                p = Porter_Stemmer_Python.PorterStemmer()
                currlist = list(self.current[j])
                self.current[j] = p.stem(currlist, 0, len(currlist) - 1)
                self.current[j] = ''.join(self.current[j])
                self.add_to_words(self.current[j])
                if j < len(self.current) - 1:
                    line = self.current[j] + " "
                else:
                    line = self.current[j] + "\n"
                self.processed.write(line)
        self.processed.close()
        self.processed = open("processed.txt", "r", encoding="utf8")
        self.add_sentence_to_words()
        self.create_excel()
        self.reduce_tdm()
        self.create_reduced()


sp = SentenceProcessor()
